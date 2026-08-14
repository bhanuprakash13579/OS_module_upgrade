"""The revenue report's three sheets must all come to the same money.

The daily report is three views of one shift: the receipts, the item table under
them, and the consolidated head-wise summary. An officer signs all three, so they
have to agree — including on the rare occasion a total is corrected by hand.

Run: python3 backend/tests/test_revenue_report.py
"""
import os, sys, tempfile
os.environ["COPS_DB_PATH"] = os.path.join(tempfile.mkdtemp(), "t.db")
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl
from types import SimpleNamespace
from app.api.duty_report import _build_revenue_sheet, _build_consolidated_sheet

FIELDS = ["baggage_duty","liquor_duty","cigarette_duty","cess_on_cig","sw_sc","gold_duty_bcd",
          "gold_duty_cons","sws_on_gold","silver_duty_cons","aidc_gold_silver","sws_on_silver",
          "aidc_on_liquor","redemption_fine","reexport_fine","personal_penalty","other_charges","fuel_duty"]

def entry(item, sl, **kw):
    e = SimpleNamespace(sl_no=sl, br_no=str(40000+sl), os_ref="", item_desc=item,
                        dutiable_value=0, gold_weight_gms=0, flight_no="", overrides=None,
                        is_sbi_challan=False, is_offline_br=False, remarks="")
    for f in FIELDS: setattr(e, f, 0)
    for k, v in kw.items(): setattr(e, k, v)
    e.total_duty = kw.get("total_duty", sum(getattr(e, f) for f in FIELDS))
    return e

day = [
    entry("IPHONE",              1, baggage_duty=8750),
    entry("IPHONE & USED LAPTOP",2, baggage_duty=5600),
    entry("IPHONE &  USED LAPTOP",3, baggage_duty=3500),   # the same item, typed loosely
    entry("FUEL",                4, fuel_duty=71897),      # fuel is an item and a column
    entry("",                    5, other_charges=20),     # nobody typed an item
    # a total corrected by hand: the columns say 5,000, the officer says 5,500
    entry("MOBILE PARTS",        6, personal_penalty=5000, total_duty=5500),
]
night = [entry("GOLD", 1, gold_duty_bcd=23448, aidc_gold_silver=3350)]

wb = openpyxl.Workbook()
ws = wb.active; ws.title = "DAY"
_build_revenue_sheet(ws, day, [], [], "DAY")
wsn = wb.create_sheet("NIGHT"); _build_revenue_sheet(wsn, night, [], [], "NIGHT")
wsc = wb.create_sheet("CONSOLIDATED"); _build_consolidated_sheet(wsc, day, night)

sheet_total = sum(e.total_duty for e in day)

# ── the item table below the receipts ──
hdr = next(r for r in range(2, ws.max_row+1)
           if str(ws.cell(r,2).value or "").strip().upper() == "ITEM")
items, r = {}, hdr+1
while str(ws.cell(r,2).value or "").strip().upper() not in ("TOTAL", ""):
    items[str(ws.cell(r,2).value)] = (ws.cell(r,4).value, ws.cell(r,5).value)
    r += 1
print("  item table:")
for k,(n,v) in items.items(): print(f"    {k:<24} {n} receipt(s)  ₹{v:,}")

assert round(sum(v for _, v in items.values())) == round(sheet_total), \
    f"the item table must come to the receipts' total: {sum(v for _,v in items.values())} vs {sheet_total}"
assert "IPHONE & USED LAPTOP" in items and items["IPHONE & USED LAPTOP"][0] == 2, \
    "two spellings of one item are one item"
assert "FUEL" in items, "fuel is an item in its own right"
assert "NOT SPECIFIED" in items, "a receipt with no item named is still money collected"
assert items["MOBILE PARTS"][1] == 5500, "a total corrected by hand is the one that carries"

# ── the consolidated ──
heads, r = {}, 3
while ws_val := wsc.cell(r,2).value:
    heads[str(ws_val)] = (wsc.cell(r,3).value or 0, wsc.cell(r,4).value or 0)
    if str(ws_val).upper().startswith("TOTAL"): break
    r += 1
day_heads = sum(d for k,(d,_) in heads.items() if not k.upper().startswith("TOTAL"))
print(f"\n  consolidated DAY column: ₹{day_heads:,}   DAY sheet: ₹{sheet_total:,}")
assert round(day_heads) == round(sheet_total), \
    f"the shift's collection must match its sheet: {day_heads} vs {sheet_total}"
assert "Correction to totals entered by hand" in heads, "and the correction is named, not hidden"
print("  correction row:", heads["Correction to totals entered by hand"])

# an ordinary day, with nothing corrected, keeps the old shape exactly
wb2 = openpyxl.Workbook(); w2 = wb2.active
_build_consolidated_sheet(w2, [entry("IPHONE", 1, baggage_duty=8750)], [])
labels = [w2.cell(r,2).value for r in range(3, 25) if w2.cell(r,2).value]
assert not any("Correction" in str(x) for x in labels), \
    "no correction, no extra row: an ordinary report is unchanged"
print("\n  ordinary report: unchanged (no correction row)")
print("  all checks passed")
