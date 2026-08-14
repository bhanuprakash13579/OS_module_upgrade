"""Shifts that go wrong in the ways real shifts go wrong.

Two of these were real defects when first written: an item named "=SUM(A1:A9)"
was stored as a live formula rather than as the words an officer typed, and a
receipt number typed with a trailing space was read as a second receipt, which
split the group and gave it two serial numbers.

Run: python3 backend/tests/test_revenue_report_hostile.py
"""
import os, sys, tempfile, math
os.environ["COPS_DB_PATH"] = os.path.join(tempfile.mkdtemp(), "t.db")
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl
from types import SimpleNamespace
from app.api.duty_report import _build_revenue_sheet, _build_consolidated_sheet

F = ["baggage_duty","liquor_duty","cigarette_duty","cess_on_cig","sw_sc","gold_duty_bcd",
     "gold_duty_cons","sws_on_gold","silver_duty_cons","aidc_gold_silver","sws_on_silver",
     "aidc_on_liquor","redemption_fine","reexport_fine","personal_penalty","other_charges","fuel_duty"]

def e(**kw):
    x = SimpleNamespace(sl_no=None, br_no="", os_ref="", item_desc="", dutiable_value=0,
                        gold_weight_gms=0, flight_no="", overrides=None, is_sbi_challan=False,
                        is_offline_br=False, remarks="", total_duty=0)
    for f in F: setattr(x, f, 0)
    for k, v in kw.items(): setattr(x, k, v)
    return x

trials = {
  "the same receipt typed twice, apart":
     [e(br_no="46552", item_desc="TV", baggage_duty=980, total_duty=980),
      e(br_no="46560", item_desc="LAPTOP", baggage_duty=5600, total_duty=5600),
      e(br_no="46552", item_desc="TV", baggage_duty=980, total_duty=980)],
  "a receipt of twelve items":
     [e(br_no="47000", item_desc=f"ITEM {i}", baggage_duty=100, total_duty=100) for i in range(12)],
  "two hundred distinct items":
     [e(br_no=str(50000+i), item_desc=f"THING {i}", baggage_duty=10, total_duty=10) for i in range(200)],
  "a refund, entered negative":
     [e(br_no="1", item_desc="TV", baggage_duty=-980, total_duty=-980)],
  "a very large seizure":
     [e(br_no="2", item_desc="GOLD", gold_duty_bcd=9_87_65_432, total_duty=9_87_65_432)],
  "fractions of a rupee":
     [e(br_no="3", item_desc="TV", baggage_duty=979.9999999, total_duty=979.9999999),
      e(br_no="4", item_desc="TV", baggage_duty=0.005, total_duty=0.005)],
  "a receipt number with spaces and letters":
     [e(br_no=" 46552-A ", item_desc="TV", baggage_duty=10, total_duty=10),
      e(br_no=" 46552-A ", item_desc="RADIO", baggage_duty=20, total_duty=20)],
  "a cancelled entry":
     [e(br_no="5", item_desc="TV CANCELLED", baggage_duty=0, total_duty=0),
      e(br_no="6", item_desc="TV", baggage_duty=980, total_duty=980)],
  "an item name of 300 characters":
     [e(br_no="7", item_desc="X"*300, baggage_duty=5, total_duty=5)],
  "a formula-looking item name":
     [e(br_no="8", item_desc="=SUM(A1:A9)", baggage_duty=5, total_duty=5)],
  "an item name with a newline":
     [e(br_no="9", item_desc="TV\nAND STAND", baggage_duty=5, total_duty=5)],
}

fails = []
for name, rows in trials.items():
    wb = openpyxl.Workbook(); ws = wb.active
    try:
        _build_revenue_sheet(ws, rows, [], [], "DAY")
        wsc = wb.create_sheet("C"); _build_consolidated_sheet(wsc, rows, [])
        path = os.path.join(tempfile.mkdtemp(), "r.xlsx"); wb.save(path)
        openpyxl.load_workbook(path)                       # it must reopen

        total = sum(r.total_duty for r in rows)
        hdr = next((r for r in range(2, ws.max_row+1)
                    if str(ws.cell(r,2).value or "").strip().upper() == "ITEM"), None)
        isum, r = 0.0, hdr + 1
        while str(ws.cell(r,2).value or "").strip().upper() not in ("TOTAL", ""):
            isum += ws.cell(r,5).value or 0; r += 1
        heads, r = 0.0, 3
        while wsc.cell(r,2).value and not str(wsc.cell(r,2).value).upper().startswith("TOTAL"):
            heads += wsc.cell(r,3).value or 0; r += 1
        if round(isum) != round(total): fails.append(f"{name}: item table {isum} != {total}")
        if round(heads) != round(total): fails.append(f"{name}: consolidated {heads} != {total}")
        print(f"  {name:<38} ₹{total:>14,.2f}  ok")
    except Exception as ex:
        fails.append(f"{name}: {type(ex).__name__}: {ex}")
        print(f"  {name:<38} FAILED — {type(ex).__name__}: {ex}")

print()
if fails:
    print("  PROBLEMS:")
    for f in fails: print("   -", f)
else:
    print("  every trial reconciled and the file reopened")


# The two that were once broken, checked directly.
ws = openpyxl.Workbook().active
_build_revenue_sheet(ws, [e(br_no="8", item_desc="=SUM(A1:A9)", baggage_duty=5, total_duty=5)], [], [], "DAY")
assert ws.cell(2, 4).data_type == "s", "an item name is text, never a formula"

ws2 = openpyxl.Workbook().active
_build_revenue_sheet(ws2, [e(br_no="46552", item_desc="TV", baggage_duty=10, total_duty=10),
                           e(br_no="46552 ", item_desc="RADIO", baggage_duty=20, total_duty=20)], [], [], "DAY")
assert [str(m) for m in ws2.merged_cells.ranges], "spacing must not split a receipt"
assert ws2.cell(3, 1).value is None, "one receipt carries one serial"
print("  the two former defects stay fixed")
