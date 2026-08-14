"""The report must survive a shift that is empty, half-typed, or odd."""
import os, sys, tempfile
os.environ["COPS_DB_PATH"] = os.path.join(tempfile.mkdtemp(), "t.db")
sys.path.insert(0, os.path.abspath("backend"))
import openpyxl
from types import SimpleNamespace
from app.api.duty_report import _build_revenue_sheet, _build_consolidated_sheet, _build_adc_sheet

F = ["baggage_duty","liquor_duty","cigarette_duty","cess_on_cig","sw_sc","gold_duty_bcd",
     "gold_duty_cons","sws_on_gold","silver_duty_cons","aidc_gold_silver","sws_on_silver",
     "aidc_on_liquor","redemption_fine","reexport_fine","personal_penalty","other_charges","fuel_duty"]

def e(**kw):
    x = SimpleNamespace(sl_no=None, br_no="", os_ref="", item_desc="", dutiable_value=None,
                        gold_weight_gms=None, flight_no="", overrides=None,
                        is_sbi_challan=False, is_offline_br=False, remarks="", total_duty=None)
    for f in F: setattr(x, f, None)          # every money column empty, as a fresh row is
    for k, v in kw.items(): setattr(x, k, v)
    return x

cases = {
    "an empty shift":               [],
    "one blank row":                [e()],
    "a row with no receipt number": [e(item_desc="TV", baggage_duty=980, total_duty=980)],
    "an item typed as spaces":      [e(br_no="1", item_desc="   ", baggage_duty=10, total_duty=10)],
    "several rows, no receipts":    [e(item_desc="A", total_duty=1), e(item_desc="B", total_duty=2)],
    "a very long description":      [e(br_no="2", item_desc="ASSORTED " * 12, total_duty=5)],
    "one receipt, five items":      [e(br_no="3", item_desc=f"ITEM {i}", total_duty=100) for i in range(5)],
    "negative correction":          [e(br_no="4", item_desc="TV", personal_penalty=5000, total_duty=4000)],
    "unicode in the item":          [e(br_no="5", item_desc="TÉLÉ — 4K “smart”", total_duty=7)],
}

for name, rows in cases.items():
    wb = openpyxl.Workbook(); ws = wb.active
    _build_revenue_sheet(ws, rows, [], [], "DAY")
    wsc = wb.create_sheet("C"); _build_consolidated_sheet(wsc, rows, [])
    wsa = wb.create_sheet("A"); _build_adc_sheet(wsa, rows, [], [])

    sheet_total = sum((r.total_duty or 0) for r in rows)
    hdr = next((r for r in range(2, ws.max_row+1)
                if str(ws.cell(r,2).value or "").strip().upper() == "ITEM"), None)
    item_sum = 0.0
    if hdr:
        r = hdr + 1
        while str(ws.cell(r,2).value or "").strip().upper() not in ("TOTAL", ""):
            item_sum += ws.cell(r,5).value or 0
            r += 1
    assert round(item_sum) == round(sheet_total), \
        f"{name}: item table {item_sum} != receipts {sheet_total}"

    heads = 0
    r = 3
    while wsc.cell(r,2).value and not str(wsc.cell(r,2).value).upper().startswith("TOTAL"):
        heads += wsc.cell(r,3).value or 0
        r += 1
    assert round(heads) == round(sheet_total), \
        f"{name}: consolidated {heads} != sheet {sheet_total}"

    # every width sane, and the file actually writes
    for col in range(1, 26):
        w = ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width
        assert w is None or 6 <= w <= 30, f"{name}: column {col} width {w}"
    wb.save(os.path.join(tempfile.mkdtemp(), "r.xlsx"))
    print(f"  {name:<30} receipts ₹{sheet_total:>7,.0f}  item table ₹{item_sum:>7,.0f}  consolidated ₹{heads:>7,.0f}")

print("  all edge cases passed")
