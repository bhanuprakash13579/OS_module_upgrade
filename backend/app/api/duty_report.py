"""
Duty Collection Report (DR) module API.
Routes are mounted at /dr.
Handles sessions, entries, tariffs, item types, settings, and Excel downloads.
"""
import io
import json
import logging
from collections import defaultdict
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.auth import User
from app.services.auth import get_current_active_user, verify_password
from app.models.duty_report import (
    DrTariff, DrItemType, DrSession, DrEntry, DrDrEntry, DrOsEntry, DrSettings,
    DrFormulaRule
)
from app.schemas.duty_report import (
    DrTariffCreate, DrTariffOut,
    DrItemTypeCreate, DrItemTypeOut,
    DrSessionCreate, DrSessionOut, SetChallanRequest,
    DrEntryIn, DrEntryOut,
    DrDrEntryIn, DrDrEntryOut,
    DrOsEntryIn, DrOsEntryOut,
    DrSettingsIn, DrSettingsOut,
    BulkSaveRequest, SessionWithEntries,
    ComputeDutiesRequest, ComputeDutiesResponse,
    DrFormulaRuleIn, DrFormulaRuleOut, DrFormulaRuleRevision,
    SubmitSessionRequest,
)

_log = logging.getLogger(__name__)
router = APIRouter()


# ── Duty computation (pure logic, no DB) ──────────────────────────────────────

_SKIP_BAGGAGE = {
    "GOLD", "SILVER", "LIQUOR", "OTHER", "CIGARETTE",
    "GOLD(C)", "SILVER(C)", "RE-EXPORT", "REEXPORT"
}


def compute_duties(tariff: DrTariff, item_desc: str, dutiable_value: float,
                   gold_weight_gms: float = 0) -> dict:
    desc_upper = (item_desc or "").upper().strip()
    auto = {}

    if desc_upper not in _SKIP_BAGGAGE:
        auto["baggage_duty"] = round(dutiable_value * tariff.baggage_rate, 2)
    else:
        auto["baggage_duty"] = 0.0

    if desc_upper == "LIQUOR":
        auto["liquor_duty"] = round(dutiable_value * tariff.liquor_duty_rate, 2)
        auto["aidc_on_liquor"] = round(dutiable_value * tariff.aidc_liquor_rate, 2)
    else:
        auto["liquor_duty"] = 0.0
        auto["aidc_on_liquor"] = 0.0

    if desc_upper == "GOLD":
        auto["gold_duty_bcd"] = round(dutiable_value * tariff.gold_bcd_rate, 2)
        auto["aidc_gold_silver"] = round(dutiable_value * tariff.aidc_gold_rate, 2)
    elif desc_upper == "SILVER":
        auto["gold_duty_bcd"] = round(dutiable_value * tariff.silver_bcd_rate, 2)
        auto["aidc_gold_silver"] = round(dutiable_value * tariff.aidc_silver_rate, 2)
    else:
        auto["gold_duty_bcd"] = 0.0

    if desc_upper == "GOLD(C)":
        auto["gold_duty_cons"] = round(dutiable_value * tariff.gold_cons_bcd_rate, 2)
        if desc_upper != "GOLD":
            auto["aidc_gold_silver"] = round(dutiable_value * tariff.aidc_gold_cons_rate, 2)
    else:
        auto["gold_duty_cons"] = 0.0

    if desc_upper == "SILVER(C)":
        auto["silver_duty_cons"] = round(dutiable_value * tariff.silver_cons_rate, 2)
        if desc_upper not in ("GOLD", "SILVER", "GOLD(C)"):
            auto["aidc_gold_silver"] = round(dutiable_value * tariff.aidc_silver_cons_rate, 2)
    else:
        auto["silver_duty_cons"] = 0.0

    if "aidc_gold_silver" not in auto:
        auto["aidc_gold_silver"] = 0.0

    total = sum([
        auto.get("baggage_duty", 0),
        auto.get("liquor_duty", 0),
        auto.get("gold_duty_bcd", 0),
        auto.get("gold_duty_cons", 0),
        auto.get("silver_duty_cons", 0),
        auto.get("aidc_gold_silver", 0),
        auto.get("aidc_on_liquor", 0),
    ])
    auto["total_duty"] = round(total)
    auto["auto_fields"] = list(auto.keys())
    return auto


# ── Tariff endpoints ──────────────────────────────────────────────────────────

@router.get("/tariffs", response_model=List[DrTariffOut])
def list_tariffs(db: Session = Depends(get_db)):
    return db.query(DrTariff).order_by(DrTariff.effective_from.desc()).all()


@router.post("/tariffs", response_model=DrTariffOut)
def create_tariff(body: DrTariffCreate, db: Session = Depends(get_db)):
    t = DrTariff(**body.model_dump())
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


@router.get("/tariffs/current", response_model=DrTariffOut)
def get_current_tariff(as_of: Optional[date] = None, db: Session = Depends(get_db)):
    ref = as_of or date.today()
    t = (db.query(DrTariff)
         .filter(DrTariff.effective_from <= ref)
         .order_by(DrTariff.effective_from.desc())
         .first())
    if not t:
        t = db.query(DrTariff).order_by(DrTariff.effective_from.asc()).first()
    if not t:
        raise HTTPException(status_code=404, detail="No tariff found. Please create one first.")
    return t


# ── Item types ────────────────────────────────────────────────────────────────

@router.get("/item-types", response_model=List[DrItemTypeOut])
def list_item_types(db: Session = Depends(get_db)):
    return db.query(DrItemType).order_by(
        DrItemType.is_system.desc(), DrItemType.usage_count.desc(), DrItemType.name
    ).all()


@router.post("/item-types", response_model=DrItemTypeOut)
def create_item_type(body: DrItemTypeCreate, db: Session = Depends(get_db)):
    existing = db.query(DrItemType).filter(
        DrItemType.name == body.name.upper().strip()
    ).first()
    if existing:
        return existing
    it = DrItemType(name=body.name.upper().strip(), is_system=body.is_system)
    db.add(it)
    db.commit()
    db.refresh(it)
    return it


@router.patch("/item-types/{item_id}/use")
def increment_item_type_usage(item_id: int, db: Session = Depends(get_db)):
    it = db.query(DrItemType).filter(DrItemType.id == item_id).first()
    if not it:
        raise HTTPException(status_code=404, detail="Item type not found")
    it.usage_count = (it.usage_count or 0) + 1
    db.commit()
    return {"id": it.id, "usage_count": it.usage_count}


# ── Settings ──────────────────────────────────────────────────────────────────

@router.get("/settings", response_model=DrSettingsOut)
def get_settings(db: Session = Depends(get_db)):
    s = db.query(DrSettings).filter(DrSettings.id == 1).first()
    if not s:
        s = DrSettings(id=1, greeting_recipients="Sir")
        db.add(s)
        db.commit()
        db.refresh(s)
    return s


@router.put("/settings", response_model=DrSettingsOut)
def update_settings(body: DrSettingsIn, db: Session = Depends(get_db)):
    s = db.query(DrSettings).filter(DrSettings.id == 1).first()
    if not s:
        s = DrSettings(id=1)
        db.add(s)
    s.greeting_recipients = body.greeting_recipients
    db.commit()
    db.refresh(s)
    return s


# ── Session endpoints ─────────────────────────────────────────────────────────

@router.get("/sessions", response_model=List[DrSessionOut])
def list_sessions(year: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(DrSession)
    if year:
        from sqlalchemy import extract
        q = q.filter(extract("year", DrSession.report_date) == year)
    return q.order_by(DrSession.report_date.desc(), DrSession.shift).all()


@router.get("/sessions/{session_id}", response_model=SessionWithEntries)
def get_session(session_id: int, db: Session = Depends(get_db)):
    sess = db.query(DrSession).filter(DrSession.id == session_id).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    entries = db.query(DrEntry).filter(
        DrEntry.session_id == session_id
    ).order_by(DrEntry.sort_order).all()
    dr_entries = db.query(DrDrEntry).filter(
        DrDrEntry.session_id == session_id
    ).order_by(DrDrEntry.sort_order).all()
    os_entries = db.query(DrOsEntry).filter(
        DrOsEntry.session_id == session_id
    ).order_by(DrOsEntry.sort_order).all()
    tariff = db.query(DrTariff).filter(DrTariff.id == sess.tariff_id).first() if sess.tariff_id else None

    result = SessionWithEntries.model_validate(sess)
    result.entries = [DrEntryOut.model_validate(e) for e in entries]
    result.dr_entries = [DrDrEntryOut.model_validate(e) for e in dr_entries]
    result.os_entries = [DrOsEntryOut.model_validate(e) for e in os_entries]
    result.tariff = DrTariffOut.model_validate(tariff) if tariff else None
    return result


@router.get("/sessions/by-date/{report_date}/{shift}", response_model=SessionWithEntries)
def get_session_by_date_shift(report_date: date, shift: str, db: Session = Depends(get_db)):
    sess = db.query(DrSession).filter(
        DrSession.report_date == report_date,
        DrSession.shift == shift.upper()
    ).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    return get_session(sess.id, db)


@router.post("/sessions", response_model=SessionWithEntries)
def create_session(body: DrSessionCreate, db: Session = Depends(get_db)):
    existing = db.query(DrSession).filter(
        DrSession.report_date == body.report_date,
        DrSession.shift == body.shift.upper()
    ).first()
    if existing:
        return get_session(existing.id, db)

    tariff = (db.query(DrTariff)
              .filter(DrTariff.effective_from <= body.report_date)
              .order_by(DrTariff.effective_from.desc())
              .first())
    if not tariff:
        tariff = db.query(DrTariff).order_by(DrTariff.effective_from.asc()).first()

    sess = DrSession(
        report_date=body.report_date,
        shift=body.shift.upper(),
        batch_name=body.batch_name,
        tariff_id=tariff.id if tariff else None,
        created_by=body.created_by,
    )
    db.add(sess)
    db.commit()
    db.refresh(sess)
    return get_session(sess.id, db)


@router.put("/sessions/{session_id}", response_model=DrSessionOut)
def update_session(session_id: int, body: DrSessionCreate, db: Session = Depends(get_db)):
    sess = db.query(DrSession).filter(DrSession.id == session_id).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    sess.batch_name = body.batch_name
    sess.created_by = body.created_by
    db.commit()
    db.refresh(sess)
    return sess


@router.patch("/sessions/{session_id}/challan", response_model=DrSessionOut)
def set_challan(session_id: int, body: SetChallanRequest, db: Session = Depends(get_db)):
    """Save the bank deposit challan number for a session's offline BR collections."""
    sess = db.query(DrSession).filter(DrSession.id == session_id).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    sess.challan_no = body.challan_no.strip()
    db.commit()
    db.refresh(sess)
    return sess


# ── Bulk entry save ───────────────────────────────────────────────────────────

# ── Carrying the BR ↔ O.S. linkage back to the register ───────────────────────

def _digit(ch: str) -> bool:
    return ch in "0123456789"


def parse_os_ref(raw: str):
    """The case a revenue line names, read out of whatever the officer typed.

    Almost every line carries a bare ``481/2026``, but the column is free text
    and the reports also hold ``OS No. 501/2026`` and
    ``OS 527/2025 DATED 22.07.2025`` — an older case settled now, with the date
    of the original order written alongside. Splitting on the first slash finds
    the year in one of those and nothing usable in the other.

    A short year is taken as this century — ``520/26`` is case 520 of 2026,
    which is the only thing anyone typing it means. Nothing is settled on that
    reading alone: the case still has to exist in the register, and a number
    that matches nothing is passed over.

    The trap is a date. ``22/07/2025`` in the case column would otherwise be read
    as case 7 of 2025 and quietly settle somebody else's case, so a number that
    follows a slash, dot or dash — as the middle of a date does — is not a case
    number.
    """
    c = raw or ""
    n = len(c)
    for i, ch in enumerate(c):
        if ch != "/":
            continue

        e = i
        while e > 0 and c[e - 1].isspace():
            e -= 1
        b = e
        while b > 0 and _digit(c[b - 1]):
            b -= 1
        if b == e or e - b > 5:
            continue
        if b > 0 and c[b - 1] in "/.-":
            continue                                  # part of a date

        k = i + 1
        while k < n and c[k].isspace():
            k += 1
        ys = k
        while k < n and _digit(c[k]):
            k += 1
        digits = k - ys
        if digits not in (2, 4):
            continue
        if k < n and c[k] in "/.-" and k + 1 < n and _digit(c[k + 1]):
            continue                                  # the date carries on

        year = int(c[ys:k]) + (2000 if digits == 2 else 0)
        if not 2000 <= year <= 2100:
            continue
        no = c[b:e].lstrip("0")
        if not no:
            continue
        return no, year
    return None


def _link_receipts_to_cases(db: Session, sess: DrSession) -> int:
    """Record, on the O.S. case, the receipts named against it in this shift.

    The revenue sheet is filled every day and carries both numbers on the same
    line: the baggage receipt, and the case it belongs to. The case is supposed
    to learn that through the adjudication screen, and at a shift change nobody
    has time to go back and open it — so the register never finds out, while the
    sheet has known all along. This reads the linkage and writes it across, with
    the date of the report the receipt appeared on, which is the day it was
    collected.

    It only ever adds. An entry already on the case is left exactly as it is,
    amount and all: the sheet is a second source for the association, not the
    authority on it. A case number that does not exist is passed over — a sheet
    may carry a typo, and that must not cost the office the shift's figures.
    """
    from app.models.offence import CopsMaster          # local: avoids a cycle

    report_date = sess.report_date.isoformat() if sess.report_date else ""
    if not report_date:
        return 0

    added_total, cases = 0, set()
    rows = db.query(DrEntry).filter(
        DrEntry.session_id == sess.id,
        DrEntry.br_no.isnot(None),
        DrEntry.os_ref.isnot(None),
    ).all()

    for r in rows:
        br_raw, os_ref = (r.br_no or "").strip(), (r.os_ref or "").strip()
        if not br_raw:
            continue
        parsed = parse_os_ref(os_ref)
        if not parsed:
            continue                                    # a note, a date, or "520" alone
        os_no, os_year = parsed

        # One cell may name several receipts.
        numbers = [n.strip() for n in br_raw.replace(";", ",").replace("/", ",").split(",")]
        numbers = [n for n in numbers if n]
        if not numbers:
            continue

        case = db.query(CopsMaster).filter(
            CopsMaster.os_no == os_no,
            CopsMaster.os_year == os_year,
            CopsMaster.entry_deleted == "N",
        ).first()
        if not case:
            continue

        try:
            existing = json.loads(case.post_adj_br_entries) if case.post_adj_br_entries else []
            if not isinstance(existing, list):
                existing = []
        except (ValueError, TypeError):
            continue                                    # unreadable: leave it alone

        added = 0
        for n in numbers:
            if any(str(e.get("no", "")).strip() == n for e in existing if isinstance(e, dict)):
                continue
            existing.append({"no": n, "date": report_date})
            added += 1
        if added:
            case.post_adj_br_entries = json.dumps(existing)
            added_total += added
            cases.add((os_no, os_year))

    if added_total:
        logging.info("revenue sheet supplied %d receipt(s) to %d case(s)",
                     added_total, len(cases))
    return len(cases)


@router.put("/sessions/{session_id}/entries", response_model=SessionWithEntries)
def save_entries(session_id: int, body: BulkSaveRequest, db: Session = Depends(get_db)):
    sess = db.query(DrSession).filter(DrSession.id == session_id).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    if sess.submitted_at:
        raise HTTPException(
            status_code=409,
            detail="Session has been submitted. Unsubmit it first before making changes."
        )

    db.query(DrEntry).filter(DrEntry.session_id == session_id).delete()
    db.query(DrDrEntry).filter(DrDrEntry.session_id == session_id).delete()
    db.query(DrOsEntry).filter(DrOsEntry.session_id == session_id).delete()

    for e in body.entries:
        db.add(DrEntry(session_id=session_id, **e.model_dump()))
    for e in body.dr_entries:
        db.add(DrDrEntry(session_id=session_id, **e.model_dump()))
    for e in body.os_entries:
        db.add(DrOsEntry(session_id=session_id, **e.model_dump()))

    db.commit()

    # The sheet knows which receipt settled which case; the register does not.
    # After the figures are safely committed, never before — a failure to link
    # must not cost the office the shift's work.
    try:
        if _link_receipts_to_cases(db, sess):
            db.commit()
    except Exception:
        db.rollback()
        logging.exception("could not carry the receipt linkage to the register")

    return get_session(session_id, db)


# ── Duty computation endpoint ─────────────────────────────────────────────────

@router.post("/compute-duties", response_model=ComputeDutiesResponse)
def compute_duties_endpoint(body: ComputeDutiesRequest, db: Session = Depends(get_db)):
    tariff = db.query(DrTariff).filter(DrTariff.id == body.tariff_id).first()
    if not tariff:
        raise HTTPException(status_code=404, detail="Tariff not found")
    result = compute_duties(tariff, body.item_desc, body.dutiable_value, body.gold_weight_gms)
    return ComputeDutiesResponse(**result)


# ── Formula Rules ─────────────────────────────────────────────────────────────

def _rules_in_force(db: Session, as_of: date):
    """The formula rules that were in force on a given day.

    A rule is never rewritten; changing one writes a new version carrying the
    same lineage and the date it takes effect. So a shift computes on the rules
    of its own report date — a sheet from last year, reopened and edited today,
    recomputes the way it did last year, and today's sheet uses today's formula.

    Versions dated after the day in question are ignored, and of those on or
    before it the newest wins. A rule with no lineage recorded is one of its own
    (a database that predates this, or a row written by hand).
    """
    newest: dict = {}
    for r in db.query(DrFormulaRule).order_by(DrFormulaRule.sort_order, DrFormulaRule.id).all():
        eff = r.effective_from
        if eff and eff > as_of:
            continue                                   # not yet in force that day
        key = r.lineage_id or r.id
        best = newest.get(key)
        if best is None or (eff or date.min) >= (best.effective_from or date.min):
            newest[key] = r
    return sorted(newest.values(), key=lambda r: (r.sort_order, r.id))


@router.get("/formula-rules", response_model=List[DrFormulaRuleOut])
def list_formula_rules(as_of: Optional[date] = None, db: Session = Depends(get_db)):
    """The rules as they stand today, or as they stood on `as_of`."""
    return _rules_in_force(db, as_of or date.today())


@router.get("/formula-rules/{rule_id}/history", response_model=List[DrFormulaRuleOut])
def formula_rule_history(rule_id: int, db: Session = Depends(get_db)):
    """Every version of one rule, newest first — what it was, and who changed it."""
    rule = db.get(DrFormulaRule, rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail="Formula rule not found")
    lineage = rule.lineage_id or rule.id
    return (db.query(DrFormulaRule)
              .filter(DrFormulaRule.lineage_id == lineage)
              .order_by(DrFormulaRule.effective_from.desc(), DrFormulaRule.id.desc())
              .all())


@router.post("/formula-rules/{rule_id}/revise", response_model=DrFormulaRuleOut)
def revise_formula_rule(
    rule_id: int,
    body: DrFormulaRuleRevision,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_active_user),
):
    """Put a new formula in force for one column, from today.

    The officer types their own username and password again to get here. A
    formula decides what every passenger is charged, and a menu item one click
    away from the sheet is a menu item that gets clicked by accident; asking for
    the password makes the change deliberate without sending anyone to find a
    supervisor at two in the morning.

    Nothing is overwritten. The old version stays in the table with its own
    dates, so a sheet from before today still computes the way it did.
    """
    if (body.username or "").strip().lower() != (user.user_id or "").strip().lower():
        raise HTTPException(status_code=403,
                            detail="Sign the change with your own username.")
    if not verify_password(body.password or "", user.user_pwd or ""):
        raise HTTPException(status_code=403, detail="That password is not right.")

    current = db.get(DrFormulaRule, rule_id)
    if not current:
        raise HTTPException(status_code=404, detail="Formula rule not found")

    today = date.today()
    lineage = current.lineage_id or current.id
    expression = (body.expression or "").strip()

    # A second change on the same day replaces that day's version rather than
    # stacking two rules with one date, which would leave the winner to chance.
    same_day = (db.query(DrFormulaRule)
                  .filter(DrFormulaRule.lineage_id == lineage,
                          DrFormulaRule.effective_from == today)
                  .first())
    if same_day and same_day.id != current.id:
        same_day.expression = expression
        same_day.changed_by = user.user_id
        db.commit(); db.refresh(same_day)
        return same_day
    if same_day and same_day.id == current.id and current.effective_from == today:
        current.expression = expression
        current.changed_by = user.user_id
        db.commit(); db.refresh(current)
        return current

    version = DrFormulaRule(
        sort_order=current.sort_order,
        target_column=current.target_column,
        column_label=current.column_label,
        condition_type=current.condition_type,
        condition_items=current.condition_items,
        expression=expression,
        is_active=current.is_active,
        notes=current.notes,
        lineage_id=lineage,
        effective_from=today,
        changed_by=user.user_id,
    )
    db.add(version)
    db.commit()
    db.refresh(version)
    logging.info("formula for %s revised by %s, in force from %s",
                 current.target_column, user.user_id, today)
    return version


@router.post("/formula-rules", response_model=DrFormulaRuleOut)
def create_formula_rule(body: DrFormulaRuleIn, db: Session = Depends(get_db)):
    rule = DrFormulaRule(**body.model_dump())
    rule.effective_from = rule.effective_from or date.today()
    db.add(rule)
    db.flush()
    rule.lineage_id = rule.lineage_id or rule.id      # a new rule begins its own line
    db.commit()
    db.refresh(rule)
    return rule


@router.put("/formula-rules/{rule_id}", response_model=DrFormulaRuleOut)
def update_formula_rule(rule_id: int, body: DrFormulaRuleIn, db: Session = Depends(get_db)):
    rule = db.get(DrFormulaRule, rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail="Formula rule not found")
    for k, v in body.model_dump().items():
        setattr(rule, k, v)
    db.commit()
    db.refresh(rule)
    return rule


@router.delete("/formula-rules/{rule_id}", status_code=204)
def delete_formula_rule(rule_id: int, db: Session = Depends(get_db)):
    rule = db.get(DrFormulaRule, rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail="Formula rule not found")
    db.delete(rule)
    db.commit()


@router.post("/formula-rules/reorder")
def reorder_formula_rules(order: List[int], db: Session = Depends(get_db)):
    """Receive an ordered list of rule IDs; assign sort_order 0,1,2,..."""
    for i, rule_id in enumerate(order):
        db.query(DrFormulaRule).filter(DrFormulaRule.id == rule_id).update(
            {"sort_order": i}, synchronize_session=False
        )
    db.commit()
    return {"ok": True}


# ── Session Submit ────────────────────────────────────────────────────────────

@router.post("/sessions/{session_id}/submit", response_model=DrSessionOut)
def submit_session(
    session_id: int,
    body: SubmitSessionRequest,
    db: Session = Depends(get_db),
):
    """Mark a session as submitted. Existing entries are NOT modified."""
    session = db.get(DrSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.submitted_at:
        raise HTTPException(status_code=409, detail="Session already submitted.")
    from datetime import datetime as _dt
    session.submitted_at = _dt.utcnow()
    session.submitted_by = body.submitted_by or "Officer"
    db.commit()
    db.refresh(session)
    # Load tariff for response
    tariff = db.get(DrTariff, session.tariff_id) if session.tariff_id else None
    session.tariff = tariff  # type: ignore[assignment]
    return session


@router.post("/sessions/{session_id}/unsubmit", response_model=DrSessionOut)
def unsubmit_session(session_id: int, db: Session = Depends(get_db)):
    """Revoke submit status (re-open the session for edits)."""
    session = db.get(DrSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    session.submitted_at = None
    session.submitted_by = None
    db.commit()
    db.refresh(session)
    tariff = db.get(DrTariff, session.tariff_id) if session.tariff_id else None
    session.tariff = tariff  # type: ignore[assignment]
    return session


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _fmt_date(d: date) -> str:
    return d.strftime("%d.%m.%Y")


def _fit_columns(ws, header_row: int, first_row: int, last_row: int,
                 columns, min_w: float = 7, max_w: float = 26) -> None:
    """Give each column the width its contents actually need.

    The widths used to be a fixed list, so a long item description was cut off
    while a column of five-digit receipt numbers sat half empty. This measures
    what is in the column and sizes it to that, within limits — nothing narrower
    than is readable, nothing so wide the sheet runs off the page.

    Headings are allowed to wrap, so a column only has to be as wide as the
    longest single word in its heading. The description below it wraps too, and
    those rows are left to size themselves, so text that needs a second line gets
    one instead of being clipped.
    """
    from openpyxl.utils import get_column_letter

    for col in columns:
        widest = 0
        for r in range(first_row, last_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            if isinstance(v, (int, float)):
                text = f"{round(v):,}"
            else:
                text = str(v)
                if text.startswith("="):
                    continue                       # a formula, not what is shown
            widest = max(widest, max((len(line) for line in text.splitlines()), default=0))
        head = str(ws.cell(row=header_row, column=col).value or "")
        longest_word = max((len(w) for w in head.split()), default=0)
        ws.column_dimensions[get_column_letter(col)].width = \
            max(min_w, min(max_w, max(widest, longest_word) + 2))


def _build_revenue_sheet(ws, entries: list, dr_entries: list, os_entries: list,
                         shift_label: str):
    """Write one DAY or NIGHT sheet into ws."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    thin  = Side(style="thin")
    med   = Side(style="medium")
    # Medium borders for main data table (looks noticeably thicker in Excel)
    border     = Border(left=med, right=med, top=med, bottom=med)
    # Thin borders for sub-tables
    sub_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    sbi_fill    = PatternFill("solid", fgColor="FFCCCC")
    total_fill  = PatternFill("solid", fgColor="FFF2CC")
    sub_hdr_fill = PatternFill("solid", fgColor="E2EFDA")

    HEADERS = [
        "SR. NO.", "BR NO. AND BR TYPE", "OS No.",
        "Item Description", "Total Dutiable Value",
        "GOLD Weight(gms)", "Baggage duty", "Liquor Duty",
        "Cigarettes Duty", "SW SC on Bagg/Lqr/Cig",
        "GOLD DUTY (BCD)", "Gold Duty (Cons.Rate BCD)",
        "Silv.Duty (Cons.Rate)", "SWS on GOLD",
        "AIDC on Gold/SILVER", "SWS ON SILVER",
        "Aidc on Liqr", "Redemption Fine", "Re-Export Fine",
        "Personal Penalty", "OTHER Charges", "FUEL DUTY",
        "CESS on CIG", "Total Duty", "Flight No"
    ]

    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 42

    sl_counter = 1
    prev_br = None
    for row_idx, entry in enumerate(entries, start=2):
        # Sub-row detection: same BR number as the row immediately above
        # Compared with the spacing taken off: a receipt typed "46552" on one row
        # and "46552 " on the next is one receipt, and reading it as two split the
        # group, gave it two serial numbers and left the pair unmerged.
        br_key = (entry.br_no or "").strip()
        is_sub_row = bool(br_key and br_key == prev_br)
        is_cancelled = "CANCEL" in (entry.item_desc or "").upper()

        # SL No.: blank for sub-rows (same BR = continuation)
        if is_sub_row:
            sl_val = None
        elif entry.sl_no is not None:
            sl_val = entry.sl_no
            sl_counter = max(sl_counter, entry.sl_no + 1)
        else:
            sl_val = sl_counter
            sl_counter += 1

        fill = sbi_fill if entry.is_sbi_challan else None
        row_data = [
            sl_val,
            entry.br_no,
            entry.os_ref,
            entry.item_desc,
            entry.dutiable_value or None,
            entry.gold_weight_gms or None,
            entry.baggage_duty or None,
            entry.liquor_duty or None,
            entry.cigarette_duty or None,
            entry.sw_sc or None,
            entry.gold_duty_bcd or None,
            entry.gold_duty_cons or None,
            entry.silver_duty_cons or None,
            entry.sws_on_gold or None,
            entry.aidc_gold_silver or None,
            entry.sws_on_silver or None,
            entry.aidc_on_liquor or None,
            entry.redemption_fine or None,
            entry.reexport_fine or None,
            entry.personal_penalty or None,
            entry.other_charges or None,
            entry.fuel_duty or None,
            entry.cess_on_cig or None,
            entry.total_duty or None,
            entry.flight_no,
        ]
        prev_br = br_key

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # An item written "=SUM(A1:A9)" — or anything else starting with an
            # equals sign — was stored as a live formula, so the description
            # showed a number, or an error, instead of the words the officer
            # wrote. It is text, and it is written as text.
            if isinstance(val, str) and val.startswith("="):
                cell.data_type = "s"
            cell.border = border
            if is_cancelled:
                # Strikethrough grey for cancelled entries
                cell.font = Font(size=9, strike=True, color="888888")
            elif col_idx == 2 and getattr(entry, "is_offline_br", False):
                cell.font = Font(size=9, color="FF0000", bold=True)
            else:
                cell.font = Font(size=9)
            cell.alignment = Alignment(
                vertical="center",
                horizontal="right" if col_idx >= 5 else "left",
                wrap_text=col_idx == 4,        # the item description, which runs long
            )
            if fill:
                cell.fill = fill
        # The row height is left unset so a wrapped description gets the second
        # line it needs instead of being cut off at a fixed 16 points.

    # One receipt covering three items is written as three rows. The serial and
    # the receipt number belong to the receipt, not to each item, so they are
    # clubbed across the group — as they are in the physical register, and as the
    # office writes them by hand. The BR was already merged this way; the serial
    # beside it was not, which left a column of numbers that no longer counted
    # receipts and did not line up with the number it belonged to.
    _ei = 0
    while _ei < len(entries):
        _ej = _ei + 1
        _key = lambda x: (x.br_no or "").strip()
        while _ej < len(entries) and _key(entries[_ej]) and _key(entries[_ej]) == _key(entries[_ei]):
            _ej += 1
        if _ej > _ei + 1:
            r1, r2 = _ei + 2, _ej + 1   # entry[i] → row i+2 (header is row 1)
            for _col, _h in ((1, "center"), (2, "left")):
                ws.merge_cells(start_row=r1, start_column=_col, end_row=r2, end_column=_col)
                ws.cell(row=r1, column=_col).alignment = Alignment(horizontal=_h, vertical="center")
        _ei = _ej

    data_last_row = len(entries) + 1

    # TOTAL row
    total_row = data_last_row + 1
    for col_idx in range(1, 26):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border
        cell.fill = total_fill
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(vertical="center", horizontal="right" if col_idx >= 5 else "left")
    ws.cell(row=total_row, column=4).value = "TOTAL"
    for col_idx in range(5, 25):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx).value = f"=SUM({col_letter}2:{col_letter}{data_last_row})"
    ws.row_dimensions[total_row].height = 18

    # Item summary table (starts 3 rows after total)
    summary_start = total_row + 3

    # Helper: apply sub-table header style to a cell
    def sub_hdr(row, col, label):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, size=9)
        cell.fill = sub_hdr_fill
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def sub_cell(row, col, val, right=False):
        cell = ws.cell(row=row, column=col, value=val)
        if isinstance(val, str) and val.startswith("="):
            cell.data_type = "s"                      # an item name, not a formula
        cell.font = Font(size=9)
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="right" if right else "left", vertical="center")

    # Item summary headers
    sub_hdr(summary_start, 2, "ITEM")
    sub_hdr(summary_start, 4, "NO OF BR")
    sub_hdr(summary_start, 5, "TOTAL DUTY")

    # DR sub-table headers
    sub_hdr(summary_start, 8, "DR")
    sub_hdr(summary_start, 9, "AMOUNT (IN Rs.)")
    sub_hdr(summary_start, 10, "ITEM")
    sub_hdr(summary_start, 11, "REMARKS")

    # OS sub-table headers
    sub_hdr(summary_start, 13, "OS")
    sub_hdr(summary_start, 14, "AMOUNT")
    sub_hdr(summary_start, 15, "ITEM")
    sub_hdr(summary_start, 16, "REMARKS")

    ws.row_dimensions[summary_start].height = 18

    # Unique items summary.
    #
    # This table has to come to the same figure as the receipts above it, so it
    # takes in every row that carries money — including one where nobody typed an
    # item, which used to be dropped and left the table quietly short.
    #
    # Items are whatever the officers type, and a new one appears here the day it
    # is first used. Two spellings of the same thing are the same thing: case and
    # stray spacing are ignored when grouping, and the first spelling used is the
    # one shown. It sums the row's total, so a total corrected by hand is carried
    # here as well.
    UNNAMED = "NOT SPECIFIED"
    item_totals: dict = defaultdict(float)
    item_counts: dict = defaultdict(int)
    item_names: dict = {}
    for e in entries:
        amount = e.total_duty or 0
        raw = (e.item_desc or "").strip()
        if not raw and not amount:
            continue                                  # an empty row of the sheet
        key = " ".join(raw.split()).upper() or UNNAMED
        item_names.setdefault(key, raw or UNNAMED)
        item_totals[key] += amount
        item_counts[key] += 1

    for i, (key, total_duty) in enumerate(item_totals.items()):
        item_name = item_names[key]
        r = summary_start + 1 + i
        sub_cell(r, 2, item_name)
        sub_cell(r, 4, item_counts[key], right=True)
        sub_cell(r, 5, round(total_duty), right=True)
        ws.row_dimensions[r].height = 15

    # Item totals row
    items_last = summary_start + len(item_totals)
    totals_r = items_last + 1
    for col in (2, 4, 5):
        cell = ws.cell(row=totals_r, column=col)
        cell.font = Font(bold=True, size=9)
        cell.fill = total_fill
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="right" if col != 2 else "left", vertical="center")
    ws.cell(row=totals_r, column=2).value = "TOTAL"
    ws.cell(row=totals_r, column=4).value = f"=SUM(D{summary_start+1}:D{items_last})"
    ws.cell(row=totals_r, column=5).value = f"=SUM(E{summary_start+1}:E{items_last})"

    # DR sub-table data
    for i, dr in enumerate(dr_entries):
        r = summary_start + 1 + i
        sub_cell(r, 8, dr.dr_no)
        sub_cell(r, 9, dr.amount, right=True)
        sub_cell(r, 10, dr.item_desc)
        sub_cell(r, 11, dr.remarks)
        ws.row_dimensions[r].height = 15

    dr_last = summary_start + len(dr_entries)
    for col in (8, 9):
        cell = ws.cell(row=dr_last + 1, column=col)
        cell.font = Font(bold=True, size=9)
        cell.fill = total_fill
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="right" if col == 9 else "left", vertical="center")
    ws.cell(row=dr_last + 1, column=8).value = "TOTAL"
    if dr_entries:
        ws.cell(row=dr_last + 1, column=9).value = f"=SUM(I{summary_start+1}:I{dr_last})"

    # OS sub-table data
    for i, os_e in enumerate(os_entries):
        r = summary_start + 1 + i
        sub_cell(r, 13, os_e.os_no)
        sub_cell(r, 14, os_e.amount, right=True)
        sub_cell(r, 15, os_e.item_desc)
        sub_cell(r, 16, os_e.remarks)
        ws.row_dimensions[r].height = 15

    os_last = summary_start + len(os_entries)
    for col in (13, 14):
        cell = ws.cell(row=os_last + 1, column=col)
        cell.font = Font(bold=True, size=9)
        cell.fill = total_fill
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="right" if col == 14 else "left", vertical="center")
    ws.cell(row=os_last + 1, column=13).value = "TOTAL"
    if os_entries:
        ws.cell(row=os_last + 1, column=14).value = f"=SUM(N{summary_start+1}:N{os_last})"

    # Offline vs Online BR split — helps officer verify bank challan amount
    offline_total = round(sum(e.total_duty or 0 for e in entries if e.is_offline_br))
    online_total  = round(sum(e.total_duty or 0 for e in entries if not e.is_offline_br))
    split_start = max(totals_r, dr_last + 1, os_last + 1) + 2
    split_fill  = PatternFill("solid", fgColor="EBF5FB")
    for offset, (label, amount) in enumerate([
        ("Offline BRs — Bank Challan", offline_total),
        ("Online BRs — Portal", online_total),
    ]):
        r = split_start + offset
        for col in (2, 24):   # col 2 = label, col 24 = Total Duty (aligns with duty totals)
            c = ws.cell(row=r, column=col)
            c.border = sub_border
            c.fill = split_fill
            c.font = Font(size=8, bold=(col == 2))
            c.alignment = Alignment(horizontal="right" if col == 24 else "left", vertical="center")
        ws.cell(row=r, column=2).value = label
        ws.cell(row=r, column=24).value = amount
        ws.row_dimensions[r].height = 14

    # Column widths — wider for better readability
    # Every column sized to what is in it, headings and sub-tables alike.
    _fit_columns(ws, 1, 1, ws.max_row, range(1, 26))

    ws.freeze_panes = "A2"


def _build_consolidated_sheet(ws, day_entries: list, night_entries: list):
    """Write the CONSOLIDATED summary sheet."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    thin  = Side(style="thin")
    med   = Side(style="medium")
    border     = Border(left=med, right=med, top=med, bottom=med)
    sub_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill  = PatternFill("solid", fgColor="FFF2CC")

    def _sum_field(entries, field):
        return round(sum(getattr(e, field, 0) or 0 for e in entries))

    DUTY_ROWS = [
        ("Baggage duty",                "baggage_duty"),
        ("Liquor Duty",                 "liquor_duty"),
        ("Cigarettes Duty",             "cigarette_duty"),
        ("CESS on Cigarettes",          "cess_on_cig"),
        ("SW SC on Bagg/Lqr/Cig",      "sw_sc"),
        ("Gold Duty (BCD)",             "gold_duty_bcd"),
        ("Gold Duty (Cons.Rate BCD)",   "gold_duty_cons"),
        ("SWS on Gold Duty",            "sws_on_gold"),
        ("Silv.Duty (Cons.Rate)",       "silver_duty_cons"),
        ("AIDC On Silv/Gold",           "aidc_gold_silver"),
        ("SWS On Silver",               "sws_on_silver"),
        ("Aidc on Liqr/Others",         "aidc_on_liquor"),
        ("Redemption Fine",             "redemption_fine"),
        ("Re-Export Fine",              "reexport_fine"),
        ("Personal Penalty/Bail",       "personal_penalty"),
        ("Misc./Other Charges",         "other_charges"),
        ("FUEL",                        "fuel_duty"),
    ]

    # Title row
    title_cell = ws.cell(row=1, column=1, value="Revenue Report — Consolidated (Day & Night Shifts)")
    title_cell.font = Font(bold=True, size=12)
    ws.merge_cells("A1:E1")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Column headers
    for col, label in [(1, "SR"), (2, "Description"), (3, "DAY (₹)"), (4, "NIGHT (₹)"), (5, "TOTAL (₹)")]:
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = Font(bold=True, size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 20

    # An officer may correct a row's total by hand — rare, and always deliberate.
    # The head-wise rows below are built from the individual columns, so they
    # cannot show that correction, and the shift's collection would come out here
    # as something other than what the DAY and NIGHT sheets say. The two must
    # agree: those sheets are the record of what was collected.
    #
    # So the difference is carried as a row of its own. Every head stays true to
    # its column, the column still adds up to the total, and the correction is
    # named rather than hidden inside a head that did not receive it. The row is
    # written only when there is a difference, so an ordinary day's report is
    # exactly as it was.
    def _corrections(entries):
        by_head = sum(_sum_field(entries, f) for _, f in DUTY_ROWS)
        return round(sum(e.total_duty or 0 for e in entries)) - by_head

    rows = list(DUTY_ROWS)
    day_fix, night_fix = _corrections(day_entries), _corrections(night_entries)
    if day_fix or night_fix:
        rows = rows + [("Correction to totals entered by hand", None)]

    for i, (label, field) in enumerate(rows):
        r = 3 + i
        if field is None:
            day_val, night_val = day_fix, night_fix
        else:
            day_val = _sum_field(day_entries, field)
            night_val = _sum_field(night_entries, field)
        for col in range(1, 6):
            cell = ws.cell(row=r, column=col)
            cell.border = sub_border
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col >= 3 else "center" if col == 1 else "left")
        ws.cell(row=r, column=1).value = i + 1
        ws.cell(row=r, column=2).value = label
        ws.cell(row=r, column=3).value = day_val or None
        ws.cell(row=r, column=4).value = night_val or None
        ws.cell(row=r, column=5).value = (day_val + night_val) or None
        ws.row_dimensions[r].height = 15

    total_row = 3 + len(rows)
    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.border = border
        cell.fill = total_fill
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="right" if col >= 3 else "left", vertical="center")
    ws.cell(row=total_row, column=2).value = "TOTAL DUTY"
    for col in (3, 4, 5):
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col).value = f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
    ws.row_dimensions[total_row].height = 20

    # The head descriptions are long and the money columns are not; each gets the
    # width it needs.
    _fit_columns(ws, 2, 2, ws.max_row, range(1, 6), min_w=6, max_w=36)


def _build_adc_sheet(ws, entries: list, dr_entries: list, os_entries: list):
    """Write the ADC report Sheet2 content."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    thin  = Side(style="thin")
    med   = Side(style="medium")
    border     = Border(left=med, right=med, top=med, bottom=med)
    sub_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill  = PatternFill("solid", fgColor="D9E1F2")
    sbi_fill     = PatternFill("solid", fgColor="FFCCCC")
    total_fill   = PatternFill("solid", fgColor="FFF2CC")
    sub_hdr_fill = PatternFill("solid", fgColor="E2EFDA")

    HEADERS = ["SR. NO.", "BR NO. AND BR TYPE", "Item Description",
               "Total Dutiable Value", "Total Duty", "Flight No"]

    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 36

    sl_counter = 1
    for row_idx, entry in enumerate(entries, start=2):
        fill = sbi_fill if entry.is_sbi_challan else None
        row_data = [
            entry.sl_no if entry.sl_no is not None else sl_counter,
            entry.br_no,
            entry.item_desc,
            entry.dutiable_value or None,
            entry.total_duty or None,
            entry.flight_no,
        ]
        if entry.sl_no is not None:
            sl_counter = entry.sl_no + 1
        else:
            sl_counter += 1

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # An item written "=SUM(A1:A9)" — or anything else starting with an
            # equals sign — was stored as a live formula, so the description
            # showed a number, or an error, instead of the words the officer
            # wrote. It is text, and it is written as text.
            if isinstance(val, str) and val.startswith("="):
                cell.data_type = "s"
            cell.border = border
            # BR No. column (2) red bold for offline BRs
            if col_idx == 2 and getattr(entry, "is_offline_br", False):
                cell.font = Font(size=9, color="FF0000", bold=True)
            else:
                cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col_idx in (4, 5) else "left")
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 16

    data_last_row = len(entries) + 1
    total_row = data_last_row + 1

    for col_idx in range(1, 7):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border
        cell.fill = total_fill
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(vertical="center",
                                   horizontal="right" if col_idx in (4, 5) else "left")
    ws.cell(row=total_row, column=1).value = "TOTAL"
    for col_idx in (4, 5):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx).value = f"=SUM({col_letter}2:{col_letter}{data_last_row})"
    ws.row_dimensions[total_row].height = 18

    # DR/OS sub-tables
    def sub_hdr(row, col, label):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, size=9)
        cell.fill = sub_hdr_fill
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def sub_cell(row, col, val, right=False):
        cell = ws.cell(row=row, column=col, value=val)
        if isinstance(val, str) and val.startswith("="):
            cell.data_type = "s"                      # an item name, not a formula
        cell.font = Font(size=9)
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="right" if right else "left", vertical="center")

    sub_start = total_row + 3
    sub_hdr(sub_start, 1, "DR"); sub_hdr(sub_start, 2, "AMOUNT"); sub_hdr(sub_start, 3, "ITEM"); sub_hdr(sub_start, 4, "REMARKS")
    sub_hdr(sub_start, 6, "OS"); sub_hdr(sub_start, 7, "AMOUNT"); sub_hdr(sub_start, 8, "ITEM"); sub_hdr(sub_start, 9, "REMARKS")
    ws.row_dimensions[sub_start].height = 18

    for i, dr in enumerate(dr_entries):
        r = sub_start + 1 + i
        sub_cell(r, 1, dr.dr_no); sub_cell(r, 2, dr.amount, right=True)
        sub_cell(r, 3, dr.item_desc); sub_cell(r, 4, dr.remarks)
        ws.row_dimensions[r].height = 15

    dr_last = sub_start + len(dr_entries)
    for col in (1, 2):
        cell = ws.cell(row=dr_last + 1, column=col)
        cell.font = Font(bold=True, size=9)
        cell.fill = total_fill
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="right" if col == 2 else "left", vertical="center")
    ws.cell(row=dr_last + 1, column=1).value = "TOTAL"
    if dr_entries:
        ws.cell(row=dr_last + 1, column=2).value = f"=SUM(B{sub_start+1}:B{dr_last})"

    for i, os_e in enumerate(os_entries):
        r = sub_start + 1 + i
        sub_cell(r, 6, os_e.os_no); sub_cell(r, 7, os_e.amount, right=True)
        sub_cell(r, 8, os_e.item_desc); sub_cell(r, 9, os_e.remarks)
        ws.row_dimensions[r].height = 15

    os_last = sub_start + len(os_entries)
    for col in (6, 7):
        cell = ws.cell(row=os_last + 1, column=col)
        cell.font = Font(bold=True, size=9)
        cell.fill = total_fill
        cell.border = sub_border
        cell.alignment = Alignment(horizontal="right" if col == 7 else "left", vertical="center")
    ws.cell(row=os_last + 1, column=6).value = "TOTAL"
    if os_entries:
        ws.cell(row=os_last + 1, column=7).value = f"=SUM(G{sub_start+1}:G{os_last})"

    # Sized to the contents, as on the revenue sheet.
    _fit_columns(ws, 1, 1, ws.max_row, range(1, 11), max_w=30)
    ws.freeze_panes = "A2"


# ── Download endpoints ────────────────────────────────────────────────────────

@router.get("/sessions/{session_id}/download/adc")
def download_adc_report(session_id: int, db: Session = Depends(get_db)):
    """Download ADC Report Excel for this session (single Sheet2 sheet)."""
    import openpyxl

    sess = db.query(DrSession).filter(DrSession.id == session_id).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")

    entries = db.query(DrEntry).filter(
        DrEntry.session_id == session_id
    ).order_by(DrEntry.sort_order).all()
    dr_entries = db.query(DrDrEntry).filter(
        DrDrEntry.session_id == session_id
    ).order_by(DrDrEntry.sort_order).all()
    os_entries = db.query(DrOsEntry).filter(
        DrOsEntry.session_id == session_id
    ).order_by(DrOsEntry.sort_order).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet2"
    _build_adc_sheet(ws, entries, dr_entries, os_entries)

    shift_code = "D" if sess.shift == "DAY" else "N"
    filename = f"{_fmt_date(sess.report_date)}({shift_code}) ADC REPORT.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/sessions/{session_id}/download/revenue")
def download_revenue_report(session_id: int, db: Session = Depends(get_db)):
    """
    Download full Revenue Report (3 sheets: DAY, NIGHT, CONSOLIDATED).
    session_id can be either the DAY or NIGHT session for this date —
    both sessions for that date will be included.
    """
    import openpyxl

    sess = db.query(DrSession).filter(DrSession.id == session_id).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")

    # Find both sessions for this date
    all_sessions = db.query(DrSession).filter(
        DrSession.report_date == sess.report_date
    ).all()
    day_sess = next((s for s in all_sessions if s.shift == "DAY"), None)
    night_sess = next((s for s in all_sessions if s.shift == "NIGHT"), None)

    def _load(s):
        if not s:
            return [], [], []
        entries = db.query(DrEntry).filter(
            DrEntry.session_id == s.id
        ).order_by(DrEntry.sort_order).all()
        dr_e = db.query(DrDrEntry).filter(
            DrDrEntry.session_id == s.id
        ).order_by(DrDrEntry.sort_order).all()
        os_e = db.query(DrOsEntry).filter(
            DrOsEntry.session_id == s.id
        ).order_by(DrOsEntry.sort_order).all()
        return entries, dr_e, os_e

    day_entries, day_dr, day_os = _load(day_sess)
    night_entries, night_dr, night_os = _load(night_sess)

    wb = openpyxl.Workbook()
    # Sheet 1: DAY
    ws_day = wb.active
    ws_day.title = "DAY"
    _build_revenue_sheet(ws_day, day_entries, day_dr, day_os, "DAY")

    # Sheet 2: NIGHT
    ws_night = wb.create_sheet("NIGHT")
    _build_revenue_sheet(ws_night, night_entries, night_dr, night_os, "NIGHT")

    # Sheet 3: CONSOLIDATED
    ws_con = wb.create_sheet("CONSOLIDATED")
    _build_consolidated_sheet(ws_con, day_entries, night_entries)

    filename = f"{_fmt_date(sess.report_date)} REVENUE REPORT.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_monthly_revenue_sheet(ws, sessions_data: list):
    """
    Build the monthly revenue report sheet matching the sample format.
    sessions_data: list of dicts {session, entries, os_entries}
    Online BRs (is_offline_br=False) are shown in blue — not part of bank challan.
    Offline BRs (is_offline_br=True) are in black — covered by session's challan_no.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    sess_fill = PatternFill("solid", fgColor="E2EFDA")   # light green for challan rows
    total_fill = PatternFill("solid", fgColor="FFF2CC")

    # Column headers matching the sample exactly (A=1 … AE=31)
    HDRS = [
        "Sl. No.", "Challan No.", "Date",
        "BR No.", "Item Description", "Quantity OF Gold/ Silver (in Gms)",
        "Total Value\nBaggage", "Total Value\nGold /Silver",
        "Baggage duty", "Liquor Duty", "Cig Duty",
        "SW SC on Bagg / Lqr / Cig",
        "Gold Duty (35%)", "Gold Duty (Cons.Rate)",
        "Silver Duty (35%)", "Silver Duty (Cons.Rate)",
        "SW SC on Gold Duty", "AIDC on Gold",
        "SW SC on Silver Duty", "AIDC on Silver Duty",
        "AIDC on Liqr", "RF", "Re-Export Fine",
        "Personal Penalty", "MISC. CHARGES",
        "CESS on CIG", "FUEL DUTY",
        "TOTAL",   # col AB (28)
        "DATES", "FORT 1/2", "REMARKS",
    ]
    for ci, h in enumerate(HDRS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=8)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 36

    _GOLD_ITEMS   = {"GOLD", "GOLD(C)"}
    _SILVER_ITEMS = {"SILVER", "SILVER(C)"}

    cur_row = 2
    sl_counter = 1
    grand_total = 0  # Python-computed to exclude per-session subtotal rows from col-28 sum

    for sd in sessions_data:
        sess    = sd["session"]
        entries = sd["entries"]     # DrEntry list
        os_entries = sd["os_entries"]  # DrOsEntry list (fuel duty etc.)

        if not entries and not os_entries:
            continue

        shift_code = "D" if sess.shift == "DAY" else "N"
        date_str = sess.report_date.strftime("%d.%m.%y") + f"({shift_code})"
        challan_str = sess.challan_no or ""

        first_row_of_session = cur_row
        prev_br = None
        session_header_written = False  # tracks whether challan/date row has been emitted yet

        for entry in entries:
            is_first = not session_header_written
            is_sub_row = bool(entry.br_no and entry.br_no == prev_br)
            # Only carry prev_br forward when the current row has a real BR number;
            # a blank/spacer row must not make the next real BR look like a sub-row.
            if entry.br_no:
                prev_br = entry.br_no
            desc_upper = (entry.item_desc or "").upper().strip()
            is_gold   = desc_upper in _GOLD_ITEMS
            is_silver = desc_upper in _SILVER_ITEMS

            # A: Sl. No. — blank for sub-rows (same BR, multiple items)
            if not is_sub_row:
                ws.cell(row=cur_row, column=1, value=sl_counter).font = Font(size=8)
            # B: Challan No. — only first row of session
            ws.cell(row=cur_row, column=2, value=challan_str if is_first else "").font = Font(size=8, bold=is_first)
            # C: Date — only first row of session
            ws.cell(row=cur_row, column=3, value=date_str if is_first else "").font = Font(size=8, bold=is_first)

            # D: BR No. — blue for online, black for offline
            br_cell = ws.cell(row=cur_row, column=4, value=entry.br_no)
            if entry.is_offline_br:
                br_cell.font = Font(size=8, color="000000")
            else:
                br_cell.font = Font(size=8, color="0070C0", bold=True)  # blue = online

            ws.cell(row=cur_row, column=5, value=entry.item_desc).font = Font(size=8)
            ws.cell(row=cur_row, column=6, value=entry.gold_weight_gms or None).font = Font(size=8)
            if is_gold or is_silver:
                ws.cell(row=cur_row, column=7, value=None)
                ws.cell(row=cur_row, column=8, value=entry.dutiable_value or None).font = Font(size=8)
            else:
                ws.cell(row=cur_row, column=7, value=entry.dutiable_value or None).font = Font(size=8)
                ws.cell(row=cur_row, column=8, value=None)

            ws.cell(row=cur_row, column=9,  value=entry.baggage_duty   or None).font = Font(size=8)
            ws.cell(row=cur_row, column=10, value=entry.liquor_duty     or None).font = Font(size=8)
            ws.cell(row=cur_row, column=11, value=entry.cigarette_duty  or None).font = Font(size=8)
            ws.cell(row=cur_row, column=12, value=entry.sw_sc           or None).font = Font(size=8)
            ws.cell(row=cur_row, column=13, value=entry.gold_duty_bcd   or None).font = Font(size=8)
            ws.cell(row=cur_row, column=14, value=entry.gold_duty_cons  or None).font = Font(size=8)
            ws.cell(row=cur_row, column=15, value=None)  # silver BCD — not a separate model field
            ws.cell(row=cur_row, column=16, value=entry.silver_duty_cons or None).font = Font(size=8)
            ws.cell(row=cur_row, column=17, value=entry.sws_on_gold   or None).font = Font(size=8)
            ws.cell(row=cur_row, column=18, value=entry.aidc_gold_silver if is_gold   else None).font = Font(size=8)
            ws.cell(row=cur_row, column=19, value=entry.sws_on_silver or None).font = Font(size=8)
            ws.cell(row=cur_row, column=20, value=entry.aidc_gold_silver if is_silver else None).font = Font(size=8)
            ws.cell(row=cur_row, column=21, value=entry.aidc_on_liquor   or None).font = Font(size=8)
            ws.cell(row=cur_row, column=22, value=entry.redemption_fine  or None).font = Font(size=8)
            ws.cell(row=cur_row, column=23, value=entry.reexport_fine    or None).font = Font(size=8)
            ws.cell(row=cur_row, column=24, value=entry.personal_penalty or None).font = Font(size=8)
            ws.cell(row=cur_row, column=25, value=entry.other_charges    or None).font = Font(size=8)
            ws.cell(row=cur_row, column=26, value=entry.cess_on_cig or None).font = Font(size=8)
            ws.cell(row=cur_row, column=27, value=entry.fuel_duty or None).font = Font(size=8)
            _entry_total = round(entry.total_duty) if entry.total_duty else 0
            ws.cell(row=cur_row, column=28, value=_entry_total or None).font = Font(size=8, bold=True)
            grand_total += _entry_total

            for ci in range(1, 32):
                c = ws.cell(row=cur_row, column=ci)
                c.border = border
                c.alignment = Alignment(vertical="center",
                                        horizontal="right" if ci >= 6 else "left")
                if is_first:
                    c.fill = sess_fill
            ws.row_dimensions[cur_row].height = 15
            if not is_sub_row:
                sl_counter += 1
            session_header_written = True
            cur_row += 1

        # Merge consecutive same-BR cells in column D within this session block
        _ei = 0
        while _ei < len(entries):
            _ej = _ei + 1
            while (_ej < len(entries) and entries[_ej].br_no
                   and entries[_ej].br_no == entries[_ei].br_no):
                _ej += 1
            if _ej > _ei + 1 and entries[_ei].br_no:
                r1 = first_row_of_session + _ei
                r2 = first_row_of_session + _ej - 1
                ws.merge_cells(start_row=r1, start_column=4, end_row=r2, end_column=4)
                ws.cell(row=r1, column=4).alignment = Alignment(
                    horizontal="left", vertical="center")
            _ei = _ej

        # OS entries (e.g. FUEL DUTY) — each gets its own row with amount in col 27
        for os_e in os_entries:
            is_first = not session_header_written
            ws.cell(row=cur_row, column=1, value=sl_counter).font = Font(size=8)
            ws.cell(row=cur_row, column=2, value=challan_str if is_first else "").font = Font(size=8, bold=is_first)
            ws.cell(row=cur_row, column=3, value=date_str if is_first else "").font = Font(size=8, bold=is_first)
            ws.cell(row=cur_row, column=4, value=os_e.os_no).font = Font(size=8)
            ws.cell(row=cur_row, column=5, value=os_e.item_desc).font = Font(size=8)
            _os_total = round(os_e.amount) if os_e.amount else 0
            ws.cell(row=cur_row, column=27, value=os_e.amount or None).font = Font(size=8)
            ws.cell(row=cur_row, column=28, value=_os_total or None).font = Font(size=8, bold=True)
            grand_total += _os_total
            for ci in range(1, 32):
                ws.cell(row=cur_row, column=ci).border = border
                ws.cell(row=cur_row, column=ci).alignment = Alignment(
                    vertical="center", horizontal="right" if ci >= 6 else "left")
            ws.row_dimensions[cur_row].height = 15
            sl_counter += 1
            session_header_written = True
            cur_row += 1

        # Per-session offline/online subtotal rows (helps officer verify challan amount)
        offline_sess = round(sum(e.total_duty or 0 for e in entries if e.is_offline_br))
        online_sess  = round(sum(e.total_duty or 0 for e in entries if not e.is_offline_br))
        sub_fill = PatternFill("solid", fgColor="EBF5FB")
        challan_label = challan_str if challan_str else "No Challan"
        for offset, (label, amt) in enumerate([
            (f"↳ Offline [{challan_label}]", offline_sess),
            ("↳ Online [Portal]", online_sess),
        ]):
            r = cur_row + offset
            for ci in range(1, 29):
                c = ws.cell(row=r, column=ci)
                c.border = border
                c.fill = sub_fill
                c.alignment = Alignment(vertical="center",
                                        horizontal="right" if ci >= 6 else "left")
                c.font = Font(size=7, italic=True, bold=(ci == 5))
            ws.cell(row=r, column=5).value = label
            ws.cell(row=r, column=28).value = amt if amt else None
            ws.row_dimensions[r].height = 11
        cur_row += 2

    # Grand total row
    if cur_row > 2:
        data_end = cur_row - 1
        for ci in range(1, 32):
            c = ws.cell(row=cur_row, column=ci)
            c.font = Font(bold=True, size=8)
            c.fill = total_fill
            c.border = border
            c.alignment = Alignment(vertical="center", horizontal="right" if ci >= 6 else "left")
        ws.cell(row=cur_row, column=5).value = "GRAND TOTAL"
        ws.cell(row=cur_row, column=5).alignment = Alignment(horizontal="right", vertical="center")
        # Cols 9-27: SUM formula is safe (subtotal rows have no data in those cols)
        for ci in range(9, 28):
            col_letter = get_column_letter(ci)
            ws.cell(row=cur_row, column=ci).value = f"=SUM({col_letter}2:{col_letter}{data_end})"
        # Col 28 (TOTAL): use Python-computed value to exclude per-session subtotal rows
        ws.cell(row=cur_row, column=28).value = grand_total or None
        ws.row_dimensions[cur_row].height = 18

    # Column widths
    widths = [5, 10, 12, 10, 24, 8, 10, 10, 10, 10, 8, 10,
              10, 10, 10, 10, 10, 10, 10, 10, 10, 8, 10, 10,
              10, 8, 10, 10, 10, 8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"


@router.get("/download/monthly-revenue")
def download_monthly_revenue(year: int, month: int, db: Session = Depends(get_db)):
    """
    Download the month-end Revenue Report Excel covering all sessions in the given month.
    Matches the format of the physical register / sample Excel exactly.
    """
    import openpyxl
    import calendar

    if not (1 <= month <= 12) or year < 2000:
        raise HTTPException(status_code=400, detail="Invalid year or month")

    from datetime import date as _date
    month_start = _date(year, month, 1)
    month_end   = _date(year, month, calendar.monthrange(year, month)[1])

    sessions = (
        db.query(DrSession)
        .filter(DrSession.report_date >= month_start, DrSession.report_date <= month_end)
        .order_by(DrSession.report_date, DrSession.shift)
        .all()
    )

    sessions_data = []
    for sess in sessions:
        entries = db.query(DrEntry).filter(
            DrEntry.session_id == sess.id
        ).order_by(DrEntry.sort_order).all()
        os_entries = db.query(DrOsEntry).filter(
            DrOsEntry.session_id == sess.id
        ).order_by(DrOsEntry.sort_order).all()
        sessions_data.append({"session": sess, "entries": entries, "os_entries": os_entries})

    wb = openpyxl.Workbook()
    ws = wb.active
    import calendar as _cal
    month_name = _cal.month_name[month]
    ws.title = f"{month_name[:3]} {year}"
    _build_monthly_revenue_sheet(ws, sessions_data)

    filename = f"{month_name} {year} - Monthly Revenue Report.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
