"""
APIS ↔ COPS Passenger Matching Service.

Matches passengers from an APIS Excel file against the COPS database.

Matching strategy (priority order):
  1. PASSPORT  — exact passport number match (highest confidence, catches changed passports too)
  2. DOB_NAME  — exact date-of-birth + name token-overlap ≥ 60%

Name normalisation handles:
  • First/last name order swaps    (SMITH JOHN  ↔  JOHN SMITH)
  • Extra tokens in one name        (JOHN FATHER SMITH  ↔  JOHN SMITH)
  • Common Indian prefixes/titles   (MR, MRS, DR, SHRI, SMT …)
  • Punctuation / non-alpha chars

Both passport match AND dob+name match are collected so a passenger who
changed passports (new passport in APIS, old passport in COPS) is still
found via the DOB+name route.
"""

import io
import re
from datetime import datetime, date
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from app.models.offence import CopsMaster, CopsItems

# Minimum token-overlap fraction to accept a DOB+name match
_NAME_THRESHOLD = 0.60


# ── Name helpers ──────────────────────────────────────────────────────────────

def _normalize_tokens(name: str) -> set[str]:
    """
    Return a set of uppercase word-tokens after stripping noise.
    Removes common titles and non-alpha characters so that
    "MR. JOHN SMITH" and "JOHN SMITH" produce identical token sets.
    """
    if not name:
        return set()
    s = name.upper()
    # Strip common honorifics / prefixes
    s = re.sub(r'\b(MR|MRS|MS|MISS|DR|PROF|SHRI|SMT|KUM|LATE|SRI)\b\.?', ' ', s)
    # Keep only letters and spaces
    s = re.sub(r'[^A-Z ]', ' ', s)
    return {t for t in s.split() if len(t) >= 2}


def _name_score(apis_name: str, cops_name: str) -> float:
    """
    Token-overlap score 0.0–1.0.

    Score = |intersection| / |shorter token-set|

    Examples:
      "JOHN SMITH"           vs "SMITH JOHN"            → 2/2 = 1.0  ✓ (reversed)
      "JOHN FATHER SMITH"    vs "JOHN SMITH"             → 2/2 = 1.0  ✓ (extra token)
      "DHANAGOPALAN K"       vs "KALYANASUNDARAM D"      → 0/1 = 0.0  ✗ (abbreviation)
      "DHANAGOPALAN KUMAR"   vs "DHANAGOPALAN KALYAN"    → 1/2 = 0.5  borderline
    """
    t1 = _normalize_tokens(apis_name)
    t2 = _normalize_tokens(cops_name)
    if not t1 or not t2:
        return 0.0
    intersection = t1 & t2
    shorter = min(len(t1), len(t2))
    return len(intersection) / shorter if shorter else 0.0


# ── Date helpers ──────────────────────────────────────────────────────────────

def _parse_dob(raw) -> Optional[date]:
    """
    Parse a DOB value that may be a Python date/datetime, a DD/MM/YYYY string,
    or an openpyxl-read cell value.
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── Excel reader ──────────────────────────────────────────────────────────────

_KNOWN_HEADERS = {'Name', 'Passport No.', 'Date of Birth', 'S.No.', 'Flight No.', 'Nationality'}


def _read_apis_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse the APIS Excel file.
    Auto-detects the header row by scanning the first 5 rows for known column names.
    Handles files with or without a title row above the headers.
    Returns a list of row dicts keyed by the header names.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Find the header row: first row (within first 5) that contains ≥2 known column names
    header_idx = None
    for i, row in enumerate(rows[:5]):
        row_vals = {str(c).strip() for c in row if c is not None}
        if len(row_vals & _KNOWN_HEADERS) >= 2:
            header_idx = i
            break

    if header_idx is None:
        return []

    raw_headers = rows[header_idx]
    headers = [str(h).strip() if h is not None else f'col_{i}'
               for i, h in enumerate(raw_headers)]

    passengers = []
    for row in rows[header_idx + 1:]:
        if not any(cell is not None for cell in row):
            continue  # skip completely empty rows
        rec = {headers[i]: row[i]
               for i in range(min(len(headers), len(row)))}
        passengers.append(rec)

    return passengers


# ── Main matching function ────────────────────────────────────────────────────

def match_from_excel(file_bytes: bytes, db: Session) -> dict:
    """
    Main entry point.

    1. Loads all active COPS masters into two in-memory indexes
       (passport → records, dob → records).
    2. Reads every APIS row.
    3. For each APIS passenger tries passport match first, then DOB+name.
    4. For every matched COPS record, fetches seized items.
    5. Returns a summary dict ready to be JSON-serialised.
    """

    # ── Build COPS indexes ────────────────────────────────────────────────────
    cops_rows = db.query(
        CopsMaster.id,
        CopsMaster.os_no,
        CopsMaster.os_year,
        CopsMaster.os_date,
        CopsMaster.location_code,
        CopsMaster.pax_name,
        CopsMaster.passport_no,
        CopsMaster.pax_date_of_birth,
        CopsMaster.pax_nationality,
        CopsMaster.total_items_value,
        CopsMaster.total_duty_amount,
        CopsMaster.total_payable,
        CopsMaster.adjudication_date,
        CopsMaster.adj_offr_name,
        CopsMaster.entry_deleted,
    ).filter(
        CopsMaster.entry_deleted != 'Y'
    ).all()

    # passport_index: normalised UPPERCASE passport → [CopsMaster row, ...]
    passport_index: dict[str, list] = {}
    # dob_index: date → [CopsMaster row, ...]
    dob_index: dict[date, list] = {}

    for r in cops_rows:
        if r.passport_no:
            key = r.passport_no.strip().upper()
            passport_index.setdefault(key, []).append(r)
        if r.pax_date_of_birth:
            dob_index.setdefault(r.pax_date_of_birth, []).append(r)

    # ── Read APIS Excel ───────────────────────────────────────────────────────
    apis_passengers = _read_apis_excel(file_bytes)
    total_apis = len(apis_passengers)

    # ── Match ─────────────────────────────────────────────────────────────────
    results: list[dict] = []

    for pax in apis_passengers:
        apis_name     = str(pax.get('Name') or '').strip()
        apis_passport = str(pax.get('Passport No.') or '').strip().upper()
        apis_dob_raw  = pax.get('Date of Birth')
        apis_dob      = _parse_dob(apis_dob_raw)
        apis_dob_str  = str(apis_dob_raw or '').strip()

        apis_flight      = str(pax.get('Flight No.')            or '').strip()
        apis_sched_date  = str(pax.get('Schedule Date')         or '').strip()
        apis_gender      = str(pax.get('Gender')                or '').strip()
        apis_nationality = str(pax.get('Nationality')           or '').strip()
        apis_sno         = pax.get('S.No.')
        apis_pnr         = str(pax.get('PNR No.')               or '').strip()
        apis_route       = str(pax.get('Embark-Disembark Port') or '').strip()

        matched_cops: list[dict] = []
        seen_ids: set[int] = set()

        # ── Priority 1: Passport exact match ─────────────────────────────────
        if apis_passport:
            for r in passport_index.get(apis_passport, []):
                if r.id in seen_ids:
                    continue
                seen_ids.add(r.id)
                matched_cops.append(_build_match(r, 'PASSPORT', 1.0))

        # ── Priority 2: DOB + name token-overlap ─────────────────────────────
        # Runs even when a passport match exists — catches cases where the
        # same person appears under a different passport number in COPS.
        if apis_dob:
            for r in dob_index.get(apis_dob, []):
                if r.id in seen_ids:
                    continue
                score = _name_score(apis_name, r.pax_name or '')
                if score >= _NAME_THRESHOLD:
                    seen_ids.add(r.id)
                    matched_cops.append(_build_match(r, 'DOB_NAME', score))

        if matched_cops:
            # Fetch seized items for every matched case
            for match in matched_cops:
                match['items'] = _fetch_items(db, match['os_no'], match['os_year'])

            results.append({
                'sno':              apis_sno,
                'apis_name':        apis_name,
                'apis_passport':    apis_passport,
                'apis_dob':         apis_dob_str,
                'apis_flight':      apis_flight,
                'apis_sched_date':  apis_sched_date,
                'apis_gender':      apis_gender,
                'apis_nationality': apis_nationality,
                'apis_pnr':         apis_pnr,
                'apis_route':       apis_route,
                'case_count':       len(matched_cops),
                'cops_matches':     matched_cops,
            })

    total_cases = sum(r['case_count'] for r in results)

    return {
        'total_apis_passengers': total_apis,
        'matched_passengers':    len(results),
        'total_cases_found':     total_cases,
        'results':               results,
    }


def _build_match(r, match_type: str, score: float) -> dict:
    return {
        'cops_id':          r.id,
        'cops_name':        r.pax_name or '',
        'cops_passport':    r.passport_no or '',
        'cops_dob':         r.pax_date_of_birth.isoformat() if r.pax_date_of_birth else '',
        'cops_nationality': r.pax_nationality or '',
        'match_type':       match_type,
        'match_score':      round(score, 2),
        'os_no':            r.os_no or '',
        'os_year':          r.os_year or 0,
        'os_date':          r.os_date.isoformat() if r.os_date else '',
        'location_code':    r.location_code or '',
        'total_items_value': float(r.total_items_value or 0),
        'total_duty_amount': float(r.total_duty_amount or 0),
        'total_payable':     float(r.total_payable or 0),
        'adjudication_date': r.adjudication_date.isoformat() if r.adjudication_date else '',
        'adj_offr_name':     r.adj_offr_name or '',
    }


def _fetch_items(db: Session, os_no: str, os_year: int) -> list[dict]:
    rows = db.query(
        CopsItems.items_sno,
        CopsItems.items_desc,
        CopsItems.items_qty,
        CopsItems.items_uqc,
        CopsItems.items_value,
        CopsItems.items_duty,
    ).filter(
        CopsItems.os_no == os_no,
        CopsItems.os_year == os_year,
        CopsItems.entry_deleted != 'Y',
    ).order_by(CopsItems.items_sno).all()

    return [
        {
            'sno':   it.items_sno or 0,
            'desc':  it.items_desc or '',
            'qty':   float(it.items_qty or 0),
            'uqc':   it.items_uqc or '',
            'value': float(it.items_value or 0),
            'duty':  float(it.items_duty or 0) if isinstance(it.items_duty, (int, float)) else 0.0,
        }
        for it in rows
    ]


# ── Excel export helper ───────────────────────────────────────────────────────

def export_to_excel(match_result: dict) -> bytes:
    """
    Convert the match_result dict into a formatted .xlsx file.
    Returns raw bytes ready to stream as a file download.

    Layout:
      Row 1  — Report title
      Row 2  — Summary stats
      Row 3  — blank
      Row 4  — Column headers
      Row 5+ — One row per (APIS passenger × matched COPS case)
               If a passenger has 3 cases they appear on 3 rows.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "APIS-COPS Matches"

    # ── Colours ───────────────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")
    PASS_FILL     = PatternFill("solid", fgColor="D1FAE5")   # green tint
    DOB_FILL      = PatternFill("solid", fgColor="FEF3C7")   # amber tint
    TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    HEADER_FONT   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    BODY_FONT     = Font(name="Calibri", size=10)
    BOLD_FONT     = Font(name="Calibri", bold=True, size=10)
    thin          = Side(style="thin", color="CCCCCC")
    BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER        = Alignment(horizontal="center", vertical="center")
    LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    COLS = [
        ("S.No.",               8),
        ("APIS Name",          28),
        ("APIS Passport",      14),
        ("APIS DOB",           12),
        ("APIS Flight",        12),
        ("APIS Date",          16),
        ("APIS Nationality",   14),
        ("Match Type",         14),
        ("COPS Name",          28),
        ("COPS Passport",      14),
        ("COPS DOB",           12),
        ("OS No.",             12),
        ("OS Date",            12),
        ("Items Value (₹)",    16),
        ("Duty (₹)",           14),
        ("Total Payable (₹)",  16),
        ("Adjudication Date",  16),
        ("Adjudicating Officer", 22),
        ("Seized Items",       50),
    ]

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"COPS ↔ APIS Match Report  |  "
        f"{match_result['matched_passengers']} matches from "
        f"{match_result['total_apis_passengers']} passengers  |  "
        f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_num = 3
    for pax in match_result['results']:
        for match in pax['cops_matches']:
            items_str = "; ".join(
                f"{it['desc']} x{it['qty']} {it['uqc']} = ₹{it['value']:,.0f}"
                for it in match.get('items', [])
                if it['desc']
            )
            fill = PASS_FILL if match['match_type'] == 'PASSPORT' else DOB_FILL

            row_data = [
                pax['sno'],
                pax['apis_name'],
                pax['apis_passport'],
                pax['apis_dob'],
                pax['apis_flight'],
                pax['apis_sched_date'],
                pax['apis_nationality'],
                match['match_type'],
                match['cops_name'],
                match['cops_passport'],
                match['cops_dob'],
                match['os_no'],
                match['os_date'],
                match['total_items_value'],
                match['total_duty_amount'],
                match['total_payable'],
                match['adjudication_date'],
                match['adj_offr_name'],
                items_str,
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font   = BODY_FONT
                cell.border = BORDER
                cell.fill   = fill
                cell.alignment = LEFT
                # Right-align numeric columns
                if col_idx in (14, 15, 16):
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            ws.row_dimensions[row_num].height = 15
            row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
